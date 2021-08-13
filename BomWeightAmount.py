import os
from xlrd import open_workbook  # for xls file
from openpyxl import load_workbook
import util

util.running_prerequisite()
logger = util.get_logger(__file__)
s, config = util.load_config()


class BomWeightAmount:

    # 設定初始變數的值，記得這些值都是要寫到EXCEL上的，
    # 先設0的好處是如果找不到相關檔案，寫上的就是0
    limit = 0
    total_weight = 0
    total_area = 0
    exclude_type = config["exclude_type"].split(',')

    def __init__(self, bom_path, bom_list_file):

        self.bom_path = bom_path  # 集中放BOM的路徑
        self.file_path = bom_list_file  # 待測試的檔案
        self.wb = load_workbook(bom_list_file, read_only=False)

    # 重整格式、清掉舊DATA
    def reformatting(self):
        table = self._get_worksheet()
        table.delete_cols(2, 10)
        table.cell(row=1, column=2, value="總重量")
        table.cell(row=1, column=3, value="總面積")
        self.save()

    # 把檔案裡面的BOM字首、編號還有一些待寫入的位置 yield出來
    def parse_bom(self):
        table = self._get_worksheet()
        for row_index, row in enumerate(table.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True)):
            if row[0] is not None:
                yield str(row[0])[0], str(row[0]), 'B{}'.format(row_index + 2), 'C{}'.format(row_index + 2)

    # 根據CELL的LOC寫DATA
    def write_data(self, weight_loc, area_loc):
        table = self._get_worksheet()

        table[weight_loc] = self.total_weight
        table[area_loc] = self.total_area

        self.reset_bom_variable()

    # 取得BOM裡面的總重跟總面積
    def get_bom_content(self, prefix, file_name):
        path = os.path.join(self.bom_path, prefix, file_name)
        self.decide_format(path)

    # 決定要用xlrd還是openpyxl
    def decide_format(self, path):
        if os.path.isfile(path + '.xls'):
            target_wb = open_workbook(path + '.xls')
            target_sh = target_wb.sheets()[0]
            self.xlrd_method(target_sh, path)
        elif os.path.isfile(path + '.xlsx'):
            target_wb = load_workbook(path+'.xlsx', read_only=False)
            target_sh = target_wb[target_wb.sheetnames[0]]
            self.openpyxl_method(target_sh, path)
        # xls或xlsx都找不到的話就只能:
        else:
            self.total_weight = '找不到該BOM檔案'
            self.total_area = '找不到該BOM檔案'
            logger.info("該路徑沒有相關檔案:{}".format(path))

    def xlrd_method(self, excel_sheet, path):
        for row in range(1, excel_sheet.nrows):
            self.limit += 1
            row = excel_sheet.row_values(row)
            if type(row[0]) != str and row[7] not in self.exclude_type:
                try:
                    self.total_weight += float(row[4]) * float(row[5])
                    self.total_area += float(row[4]) * float(row[6])
                except:
                    logger.info("{},BOM檔案重量或面積異常，:{}".format(path, [type(row[5]), row[5], type(row[6]), row[6]]))
                    self.total_weight = 'BOM異常'
                    self.total_area = 'BOM異常'
                    return False
            if self.limit == 100:
                break

    def openpyxl_method(self, excel_sheet, path):
        for row in excel_sheet.iter_rows(min_row=3, min_col=1, max_col=12, values_only=True):
            self.limit += 1
            if row[0] is not None and row[7] not in self.exclude_type:
                try:
                    self.total_weight += float(row[4]) * float(row[5])
                    self.total_area += float(row[4]) * float(row[6])
                except TypeError as e:
                    logger.info("{},BOM檔案異常:{}{}".format(path, e, [type(row[5]), row[5], type(row[6]), row[6]]))
                    self.total_weight = 'BOM異常'
                    self.total_area = 'BOM異常'
                    return False

            if self.limit == 100:
                break

    # 重置變數值
    def reset_bom_variable(self):
        self.total_weight = 0
        self.total_area = 0
        self.limit = 0

    def _get_worksheet(self):
        return self.wb[self.wb.sheetnames[0]]

    def save(self):
        self.wb.save(self.file_path)
